
# Create your tests here.
from django.test import TestCase, Client
from django.urls import reverse
from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from users.models import StudentProfile
from .models import Course, Grade
from .utils import save_grade, get_letter_grade, determine_eligibility
import openpyxl
from io import BytesIO

class CourseModelTests(TestCase):
    def test_course_creation(self):
        """Test creating a course with valid data"""
        course = Course.objects.create(code="TC101", name="Test Course")
        self.assertEqual(course.code, "TC101")
        self.assertEqual(course.name, "Test Course")
        self.assertEqual(str(course), "TC101")

    def test_code_unique_constraint(self):
        """Test that course code is unique"""
        Course.objects.create(code="TC101", name="Test Course")
        with self.assertRaises(Exception):
            Course.objects.create(code="TC101", name="Another Course")

    def test_code_max_length(self):
        """Test that code respects max_length constraint"""
        long_code = "x" * 11
        with self.assertRaises(Exception):
            Course.objects.create(code=long_code, name="Test Course")

    def test_name_max_length(self):
        """Test that name respects max_length constraint"""
        long_name = "x" * 101
        with self.assertRaises(Exception):
            Course.objects.create(code="TC102", name=long_name)

class GradeModelTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='student', password='studentpass123')
        self.student = StudentProfile.objects.create(user=self.user)
        self.course = Course.objects.create(code="TC101", name="Test Course")

    def test_grade_creation(self):
        """Test creating a grade with valid data"""
        grade = Grade.objects.create(
            student=self.student,
            course=self.course,
            midterm_grade=80.0,
            final_exam_grade=85.0,
            final_grade=83.0,
            letter_grade="BB",
            eligibility="Not Eligible",
            absences=2,
            declared_resit=False
        )
        self.assertEqual(grade.student, self.student)
        self.assertEqual(grade.course, self.course)
        self.assertEqual(grade.midterm_grade, 80.0)
        self.assertEqual(grade.letter_grade, "BB")
        self.assertEqual(str(grade), f"{self.user.email} - TC101 - BB")

    def test_absences_default(self):
        """Test that absences defaults to 0"""
        grade = Grade.objects.create(
            student=self.student,
            course=self.course,
            eligibility="Not Eligible"
        )
        self.assertEqual(grade.absences, 0)

    def test_declared_resit_default(self):
        """Test that declared_resit defaults to False"""
        grade = Grade.objects.create(
            student=self.student,
            course=self.course,
            eligibility="Not Eligible"
        )
        self.assertFalse(grade.declared_resit)

class UtilsTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='student', password='studentpass123')
        self.student = StudentProfile.objects.create(user=self.user)
        self.course = Course.objects.create(code="TC101", name="Test Course")

    def test_get_letter_grade(self):
        """Test get_letter_grade function"""
        self.assertEqual(get_letter_grade(95), "AA")
        self.assertEqual(get_letter_grade(87), "BA")
        self.assertEqual(get_letter_grade(82), "BB")
        self.assertEqual(get_letter_grade(77), "CB")
        self.assertEqual(get_letter_grade(72), "CC")
        self.assertEqual(get_letter_grade(67), "DC")
        self.assertEqual(get_letter_grade(62), "DD")
        self.assertEqual(get_letter_grade(57), "FD")
        self.assertEqual(get_letter_grade(50), "FF")

    def test_determine_eligibility(self):
        """Test determine_eligibility function"""
        self.assertEqual(determine_eligibility("DD"), "Eligible")
        self.assertEqual(determine_eligibility("FD"), "Eligible")
        self.assertEqual(determine_eligibility("FF"), "Eligible")
        self.assertEqual(determine_eligibility("CC"), "Not Eligible")
        self.assertEqual(determine_eligibility("AA"), "Not Eligible")

    def test_save_grade(self):
        """Test save_grade function"""
        save_grade(self.student, self.course, 80, 85, 2)
        grade = Grade.objects.get(student=self.student, course=self.course)
        self.assertEqual(grade.midterm_grade, 80)
        self.assertEqual(grade.final_exam_grade, 85)
        self.assertEqual(grade.final_grade, 83.0)  # 80*0.4 + 85*0.6
        self.assertEqual(grade.letter_grade, "BB")
        self.assertEqual(grade.eligibility, "Not Eligible")
        self.assertEqual(grade.absences, 2)

    def test_save_grade_with_excessive_absences(self):
        """Test save_grade with absences > 3"""
        save_grade(self.student, self.course, 80, 85, 4)
        grade = Grade.objects.get(student=self.student, course=self.course)
        self.assertEqual(grade.letter_grade, "DZ")
        self.assertEqual(grade.eligibility, "Not Eligible")

class ExamsViewTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='student', password='studentpass123', email='student@example.com')
        self.student = StudentProfile.objects.create(user=self.user)
        self.course = Course.objects.create(code="TC101", name="Test Course")
        self.grade = Grade.objects.create(
            student=self.student,
            course=self.course,
            midterm_grade=60,
            final_exam_grade=60,
            final_grade=60,
            letter_grade="DD",
            eligibility="Eligible",
            absences=0,
            declared_resit=True
        )

    def test_studentgrade_view(self):
        """Test studentgrade view for authenticated student"""
        self.client.login(username='student', password='studentpass123')
        response = self.client.get(reverse('studentgrade'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'studentgrade.html')
        self.assertIn('original_grades', response.context)
        self.assertIn('resit_grades', response.context)
        self.assertIn('gpa', response.context)
        self.assertEqual(len(response.context['original_grades']), 1)
        self.assertEqual(response.context['gpa'], 1.0)  # DD = 1.0 GPA

    def test_studentgrade_view_no_profile(self):
        """Test studentgrade view redirects for user without student profile"""
        user = User.objects.create_user(username='noprofile', password='noprofile123')
        self.client.login(username='noprofile', password='noprofile123')
        response = self.client.get(reverse('studentgrade'))
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('some_error_page'))

    def test_insexam_view(self):
        """Test insexam view with pagination and filtering"""
        response = self.client.get(reverse('insexam'), {'course': 'TC101', 'eligibility': 'Eligible'})
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'insexam.html')
        self.assertIn('students', response.context)
        self.assertIn('page_obj', response.context)
        self.assertIn('courses', response.context)
        self.assertEqual(len(response.context['students']), 1)
        self.assertEqual(response.context['selected_course'], 'TC101')
        self.assertEqual(response.context['selected_eligibility'], 'Eligible')

    def test_resitgrades_view(self):
        """Test resitgrades view renders correctly"""
        response = self.client.get(reverse('resitgrades'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'resitgrades.html')

    def test_declare_resit_success(self):
        """Test declare_resit view with valid data"""
        self.client.login(username='student', password='studentpass123')
        response = self.client.post(
            reverse('declare_resit'),
            {'course_code': 'TC101'},
            content_type='application/x-www-form-urlencoded'
        )
        self.assertEqual(response.status_code, 200)
        self.assertJSONEqual(response.content, {'status': 'success'})
        self.assertTrue(Grade.objects.get(student=self.student, course=self.course).declared_resit)

    def test_declare_resit_invalid_course(self):
        """Test declare_resit with invalid course code"""
        self.client.login(username='student', password='studentpass123')
        response = self.client.post(
            reverse('declare_resit'),
            {'course_code': 'INVALID'},
            content_type='application/x-www-form-urlencoded'
        )
        self.assertEqual(response.status_code, 404)
        self.assertJSONEqual(response.content, {'status': 'error', 'message': 'No eligible resit found'})

    def test_delete_grade_success(self):
        """Test delete_grade view with valid grade ID"""
        response = self.client.post(
            reverse('delete_grade'),
            {'id': self.grade.id},
            content_type='application/x-www-form-urlencoded'
        )
        self.assertEqual(response.status_code, 200)
        self.assertJSONEqual(response.content, {'status': 'success', 'message': 'Deleted successfully'})
        self.assertFalse(Grade.objects.filter(id=self.grade.id).exists())

    def test_delete_grade_invalid_id(self):
        """Test delete_grade with invalid grade ID"""
        response = self.client.post(
            reverse('delete_grade'),
            {'id': 999},
            content_type='application/x-www-form-urlencoded'
        )
        self.assertEqual(response.status_code, 404)
        self.assertJSONEqual(response.content, {'status': 'error', 'message': 'Record not found'})

    def test_upload_grades_regular_valid(self):
        """Test upload_grades with valid regular Excel file"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['email', 'midterm', 'final_exam', 'absences'])
        sheet.append(['student@example.com', 80, 85, 2])
        excel_content = BytesIO()
        workbook.save(excel_content)
        excel_content.seek(0)
        uploaded_file = SimpleUploadedFile(
            "test.xlsx",
            excel_content.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response = self.client.post(
            reverse('upload_grades', args=[self.course.id]),
            {'grade_file': uploaded_file, 'upload_type': 'regular'},
            format='multipart'
        )
        self.assertEqual(response.status_code, 200)
        self.assertJSONEqual(response.content, {'status': 'success'})
        grade = Grade.objects.get(student=self.student, course=self.course)
        self.assertEqual(grade.midterm_grade, 80)
        self.assertEqual(grade.final_exam_grade, 85)
        self.assertEqual(grade.absences, 2)
        self.assertEqual(grade.letter_grade, "BB")

    def test_upload_grades_resit_valid(self):
        """Test upload_grades with valid resit Excel file"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['email', 'resit_grade'])
        sheet.append(['student@example.com', 90])
        excel_content = BytesIO()
        workbook.save(excel_content)
        excel_content.seek(0)
        uploaded_file = SimpleUploadedFile(
            "test.xlsx",
            excel_content.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response = self.client.post(
            reverse('upload_grades', args=[self.course.id]),
            {'grade_file': uploaded_file, 'upload_type': 'resit'},
            format='multipart'
        )
        self.assertEqual(response.status_code, 200)
        self.assertJSONEqual(response.content, {'status': 'success'})
        grade = Grade.objects.get(student=self.student, course=self.course)
        self.assertEqual(grade.resit_exam_grade, 90)

    def test_upload_grades_invalid_data(self):
        """Test upload_grades with invalid Excel data"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['email', 'midterm', 'final_exam'])
        sheet.append(['student@example.com', 80])  # Missing final_exam
        excel_content = BytesIO()
        workbook.save(excel_content)
        excel_content.seek(0)
        uploaded_file = SimpleUploadedFile(
            "test.xlsx",
            excel_content.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response = self.client.post(
            reverse('upload_grades', args=[self.course.id]),
            {'grade_file': uploaded_file, 'upload_type': 'regular'},
            format='multipart'
        )
        self.assertEqual(response.status_code, 500)
        self.assertIn('Incomplete data', response.json()['message'])

    def test_download_resit_excel(self):
        """Test download_resit_excel view"""
        response = self.client.get(reverse('download_resit_excel'), {'course_code': 'TC101'})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertEqual(response['Content-Disposition'], 'attachment; filename="TC101_resit_students.xlsx"')
        # Verify Excel content
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        sheet = workbook.active
        self.assertEqual(sheet['A1'].value, 'Student Email')
        self.assertEqual(sheet['A2'].value, 'student@example.com')
