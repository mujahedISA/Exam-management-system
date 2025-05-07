from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse, HttpResponseNotAllowed
import openpyxl
from django.core.paginator import Paginator
from exams.utils import determine_eligibility, get_letter_grade, save_grade
from .models import Grade, StudentProfile, Course
from django.views.decorators.http import require_POST


def declare_resit(request):
    if request.method == 'POST':
        course_code = request.POST.get('course_code')
        
        if not hasattr(request.user, 'studentprofile'):
            return JsonResponse({'status': 'error', 'message': 'Student profile not found'}, status=403)

        student = request.user.studentprofile

        try:
            resit = Grade.objects.get(student=student, course__code=course_code)
            resit.declared_resit = True
            resit.save()
            return JsonResponse({'status': 'success'})
        except Grade.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'No eligible resit found'}, status=404)

    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)

def delete_grade(request):
    if request.method == 'POST':
        resit_id = request.POST.get('id')

        try:
            resit = Grade.objects.get(id=resit_id)
            resit.delete()
            return JsonResponse({'status': 'success', 'message': 'Deleted successfully'})
        except Grade.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Record not found'}, status=404)

    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)



@require_POST
def upload_grades(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    upload_type = request.POST.get('upload_type', 'regular')  # default to 'regular'
    excel_file = request.FILES.get('grade_file')

    if not excel_file:
        return JsonResponse({'status': 'error', 'message': 'No file uploaded'}, status=400)

    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active  # Use the first sheet

        if upload_type == "resit":
            for row in sheet.iter_rows(min_row=2, values_only=True):
                print(f"DEBUG: Row: {row}")
                if len(row) != 2:
                    raise ValueError(f"Invalid row format: {row}")

                email, resit_grade = row

                try:
                    student = StudentProfile.objects.get(user__email=email)
                except StudentProfile.DoesNotExist:
                    raise ValueError(f"StudentProfile not found for email: {email}")

                try:
                    grade_obj = Grade.objects.get(student=student, course=course)
                except Grade.DoesNotExist:
                    raise ValueError(f"Grade record not found for {email} in course {course.name}")

                grade_obj.resit_exam_grade = resit_grade
                grade_obj.save()
                print(f"CONFIRM: {email} => Saved resit grade: {grade_obj.resit_exam_grade}")

        else:
            # Regular upload: email, midterm, final, [optional absences]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                print(f"DEBUG: Row: {row}")
                if len(row) < 3:
                    raise ValueError(f"Incomplete data in row: {row}")

                email, midterm, final_exam, *optional_absences = row
                absences = optional_absences[0] if optional_absences else 0

                student = get_object_or_404(StudentProfile, user__email=email)
                save_grade(student, course, midterm, final_exam, absences)

        return JsonResponse({'status': 'success'})

    except Exception as e:
        print(f"ERROR: {e}")
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    
def download_resit_excel(request):
    course_code = request.GET.get('course_code')

    if not course_code:
        return HttpResponse("No course selected.", status=400)

    # Filter students who declared resit for the selected course
    students = Grade.objects.filter(
        course__code=course_code,
        declared_resit=True
    ).select_related('student__user')

    # Create an Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{course_code} Resits"

    # Write header
    ws.append(['Student Email'])

    # Write emails
    for student in students:
        ws.append([student.student.user.email])

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"{course_code}_resit_students.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response



def insexam(request):
    # Get all courses
    courses = Course.objects.all()

    # Separate filters for both tables
    selected_course = request.GET.get('course')
    selected_eligibility = request.GET.get('eligibility')
    selected_course2 = request.GET.get('course2')
    selected_eligibility2 = request.GET.get('eligibility2')

    # Base querysets with ordering to avoid UnorderedObjectListWarning
    students = Grade.objects.all().order_by('id')  # You can change 'id' to another field like 'student__user__email'
    resit_students = Grade.objects.filter(declared_resit=True).order_by('id')

    # Apply filters
    if selected_course:
        students = students.filter(course__code=selected_course)

    if selected_eligibility:
        students = students.filter(eligibility=selected_eligibility)

    if selected_course2:
        resit_students = resit_students.filter(course__code=selected_course2)

    if selected_eligibility2:
        resit_students = resit_students.filter(eligibility=selected_eligibility2)

    # Pagination for filtered students
    page_number = request.GET.get('page', 1)
    paginator = Paginator(students, 5)
    page_obj = paginator.get_page(page_number)

    # Pagination for filtered resit students
    page_number2 = request.GET.get('page2', 1)
    paginator2 = Paginator(resit_students, 5)
    page_obj2 = paginator2.get_page(page_number2)

    context = {
        'students': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator,
        'selected_course': selected_course,
        'selected_eligibility': selected_eligibility,

        'students2': page_obj2.object_list,
        'page_obj2': page_obj2,
        'paginator2': paginator2,
        'selected_course2': selected_course2,
        'selected_eligibility2': selected_eligibility2,

        'courses': courses,
    }

    return render(request, 'insexam.html', context)

def resitgrades(request):
    return render(request, 'resitgrades.html')



def studentgrade(request):
    if not hasattr(request.user, 'studentprofile'):
        return redirect('some_error_page')

    student = request.user.studentprofile
    grades = Grade.objects.filter(student=student)

    total_points = 0
    total_courses = 0

    letter_to_gpa = {
        'AA': 4.0, 'BA': 3.5, 'BB': 3.0,
        'CB': 2.5, 'CC': 2.0, 'DC': 1.5,
        'DD': 1.0, 'FD': 0.5, 'FF': 0.0, 'DZ': 0.0,
    }

    original_grades = []
    resit_grades = []

    for grade in grades:
        # Original calculation
        if grade.absences > 3:
            original_letter = "DZ"
            original_final = (grade.midterm_grade * 0.4) + (grade.final_exam_grade * 0.6)
        else:
            original_final = (grade.midterm_grade * 0.4) + (grade.final_exam_grade * 0.6)
            original_letter = get_letter_grade(original_final)

        grade.final_grade = original_final
        grade.letter_grade = original_letter

        # Resit calculation
        if grade.resit_exam_grade is not None:
            resit_final = (grade.midterm_grade * 0.4) + (grade.resit_exam_grade * 0.6)

            if grade.absences > 3:
                resit_letter = "DZ"
            else:
                resit_letter = get_letter_grade(resit_final)

            grade.resit_final_grade = resit_final
            grade.resit_letter_grade = resit_letter

            # ✅ Use resit grade for GPA
            gpa_point = letter_to_gpa.get(resit_letter, 0.0)
            resit_grades.append(grade)
        else:
            grade.resit_final_grade = None
            grade.resit_letter_grade = None

            # ✅ Use original grade for GPA
            gpa_point = letter_to_gpa.get(original_letter, 0.0)

        grade.save()
        total_points += gpa_point
        total_courses += 1
        original_grades.append(grade)

    gpa = round(total_points / total_courses, 2) if total_courses > 0 else None

    return render(request, 'studentgrade.html', {
        'original_grades': original_grades,
        'resit_grades': resit_grades,
        'gpa': gpa,
    })






