
# Create your tests here.

from django.test import TestCase, Client
from django.urls import reverse
from django.contrib.auth.models import User, Group
from django.core.files.uploadedfile import SimpleUploadedFile
from exams.models import Course, Grade
from users.models import StudentProfile
from .models import ResitExamSchedule, ResitExamContent
from .forms import ExcelUploadForm
import openpyxl
from io import BytesIO

class ResitExamScheduleModelTests(TestCase):
    def setUp(self):
        # Create a test course for model tests
        self.course = Course.objects.create(name="Test Course", code="TC101")

    def test_resit_exam_schedule_creation(self):
        """Test creating a ResitExamSchedule with valid data"""
        schedule = ResitExamSchedule.objects.create(
            course=self.course,
            place="Room 101",
            date="2025-06-01"
        )
        self.assertEqual(schedule.course, self.course)
        self.assertEqual(schedule.place, "Room 101")
        self.assertEqual(schedule.date, "2025-06-01")
        self.assertEqual(str(schedule), f"{self.course} - 2025-06-01")

    def test_place_max_length(self):
        """Test that place respects max_length constraint"""
        max_length = ResitExamSchedule._meta.get_field('place').max_length
        long_place = "x" * (max_length + 1)
        with self.assertRaises(Exception):
            ResitExamSchedule.objects.create(
                course=self.course,
                place=long_place,
                date="2025-06-01"
            )

    def test_date_max_length(self):
        """Test that date respects max_length constraint"""
        max_length = ResitExamSchedule._meta.get_field('date').max_length
        long_date = "x" * (max_length + 1)
        with self.assertRaises(Exception):
            ResitExamSchedule.objects.create(
                course=self.course,
                place="Room 101",
                date=long_date
            )

class ResitExamContentModelTests(TestCase):
    def setUp(self):
        # Create a test course for model tests
        self.course = Course.objects.create(name="Test Course", code="TC101")

    def test_resit_exam_content_creation(self):
        """Test creating a ResitExamContent with valid data"""
        content = ResitExamContent.objects.create(
            course=self.course,
            exam_type="Multiple Choice",
            num_questions=50,
            calculator_allowed=True,
            additional_notes="Bring pencils"
        )
        self.assertEqual(content.course, self.course)
        self.assertEqual(content.exam_type, "Multiple Choice")
        self.assertEqual(content.num_questions, 50)
        self.assertTrue(content.calculator_allowed)
        self.assertEqual(content.additional_notes, "Bring pencils")

    def test_num_questions_positive(self):
        """Test that num_questions must be positive"""
        with self.assertRaises(Exception):
            ResitExamContent.objects.create(
                course=self.course,
                exam_type="Multiple Choice",
                num_questions=-1,
                calculator_allowed=False
            )

    def test_exam_type_max_length(self):
        """Test that exam_type respects max_length constraint"""
        max_length = ResitExamContent._meta.get_field('exam_type').max_length
        long_exam_type = "x" * (max_length + 1)
        with self.assertRaises(Exception):
            ResitExamContent.objects.create(
                course=self.course,
                exam_type=long_exam_type,
                num_questions=50,
                calculator_allowed=False
            )

class ExcelUploadFormTests(TestCase):
    def test_valid_excel_file(self):
        """Test form validation with valid Excel file"""
        excel_content = BytesIO()
        workbook = openpyxl.Workbook()
        workbook.save(excel_content)
        excel_content.seek(0)
        uploaded_file = SimpleUploadedFile("test.xlsx", excel_content.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        form = ExcelUploadForm(files={'excel_file': uploaded_file})
        self.assertTrue(form.is_valid())

    def test_invalid_file_type(self):
        """Test form validation with invalid file type"""
        uploaded_file = SimpleUploadedFile("test.txt", b"not an excel file", content_type="text/plain")
        form = ExcelUploadForm(files={'excel_file': uploaded_file})
        self.assertFalse(form.is_valid())
        self.assertIn('excel_file', form.errors)

class DeclarationsViewTests(TestCase):
    def setUp(self):
        # Set up test client and users
        self.client = Client()
        self.student_user = User.objects.create_user(username='student', password='studentpass123')
        self.faculty_user = User.objects.create_user(username='faculty', password='facultypass123')
        
        # Create groups
        self.student_group = Group.objects.create(name='student')
        self.faculty_group = Group.objects.create(name='faculty')
        self.student_user.groups.add(self.student_group)
        self.faculty_user.groups.add(self.faculty_group)
        
        # Create test course and related objects
        self.course = Course.objects.create(name="Test Course", code="TC101")
        self.student_profile = StudentProfile.objects.create(user=self.student_user)
        self.grade = Grade.objects.create(
            student=self.student_profile,
            course=self.course,
            declared_resit=True
        )

    def test_resitannouncement_view_student(self):
        """Test resitannouncement view for authenticated student"""
        self.client.login(username='student', password='studentpass123')
        
        # Create test schedule and content
        schedule = ResitExamSchedule.objects.create(course=self.course, place="Room 101", date="2025-06-01")
        content = ResitExamContent.objects.create(
            course=self.course,
            exam_type="Multiple Choice",
            num_questions=50,
            calculator_allowed=True
        )
        
        response = self.client.get(reverse('resitannouncement'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'resitannouncement.html')
        self.assertIn('resit_details', response.context)
        self.assertIn(self.course.id, response.context['resit_details'])
        self.assertEqual(response.context['resit_details'][self.course.id]['schedule'], schedule)
        self.assertEqual(response.context['resit_details'][self.course.id]['content'], content)

    def test_resitannouncement_view_non_student(self):
        """Test resitannouncement view redirects for non-student"""
        self.client.login(username='faculty', password='facultypass123')
        response = self.client.get(reverse('resitannouncement'))
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('student_page'))

    def test_facultysecexam_view(self):
        """Test facultysecexam view renders correctly"""
        response = self.client.get(reverse('facultysecexam'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'facultysecexam.html')

    def test_upload_resit_schedule_unauthorized(self):
        """Test upload_resit_schedule view for non-faculty user"""
        self.client.login(username='student', password='studentpass123')
        response = self.client.post(reverse('upload_resit_schedule'))
        self.assertEqual(response.status_code, 403)
        self.assertJSONEqual(response.content, {'status': 'error', 'message': 'Unauthorized access.'})

    def test_upload_resit_schedule_valid_excel(self):
        """Test upload_resit_schedule with valid Excel file"""
        self.client.login(username='faculty', password='facultypass123')
        
        # Create Excel file
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['Course ID', 'Course Name', 'Place', 'Date'])
        sheet.append(['TC101', 'Test Course', 'Room 101', '2025-06-01'])
        excel_content = BytesIO()
        workbook.save(excel_content)
        excel_content.seek(0)
        uploaded_file = SimpleUploadedFile(
            "test.xlsx",
            excel_content.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        response = self.client.post(
            reverse('upload_resit_schedule'),
            {'excel_file': uploaded_file},
            format='multipart'
        )
        self.assertEqual(response.status_code, 200)
        self.assertJSONEqual(
            response.content,
            {'status': 'success', 'message': 'Resit exam schedule uploaded successfully.'}
        )
        self.assertTrue(ResitExamSchedule.objects.filter(course=self.course, place="Room 101").exists())

    def test_upload_resit_schedule_invalid_excel(self):
        """Test upload_resit_schedule with invalid Excel headers"""
        self.client.login(username='faculty', password='facultypass123')
        
        # Create Excel file with wrong headers
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['Wrong Header', 'Course Name', 'Place', 'Date'])
        excel_content = BytesIO()
        workbook.save(excel_content)
        excel_content.seek(0)
        uploaded_file = SimpleUploadedFile(
            "test.xlsx",
            excel_content.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        response = self.client.post(
            reverse('upload_resit_schedule'),
            {'excel_file': uploaded_file},
            format='multipart'
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn('Excel file must contain columns', response.json()['message'])

    def test_upload_resit_details_valid(self):
        """Test upload_resit_details with valid Excel file"""
        self.client.login(username='faculty', password='facultypass123')
        
        # Create Excel file
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['num_questions', 'exam_type', 'calculator_allowed', 'additional_notes'])
        sheet.append([50, 'Multiple Choice', 'Yes', 'Bring pencils'])
        excel_content = BytesIO()
        workbook.save(excel_content)
        excel_content.seek(0)
        uploaded_file = SimpleUploadedFile(
            "test.xlsx",
            excel_content.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        response = self.client.post(
            reverse('upload_resit_details', args=[self.course.id]),
            {'excel_file': uploaded_file},
            format='multipart'
        )
        self.assertEqual(response.status_code, 200)
        self.assertJSONEqual(response.content, {'status': 'success'})
        self.assertTrue(ResitExamContent.objects.filter(
            course=self.course,
            num_questions=50,
            exam_type="Multiple Choice",
            calculator_allowed=True
        ).exists())

    def test_upload_resit_details_invalid_data(self):
        """Test upload_resit_details with invalid data"""
        self.client.login(username='faculty', password='facultypass123')
        
        # Create Excel file with invalid data
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['num_questions', 'exam_type', 'calculator_allowed', 'additional_notes'])
        sheet.append(['invalid', 'Multiple Choice', 'Yes', 'Notes'])
        excel_content = BytesIO()
        workbook.save(excel_content)
        excel_content.seek(0)
        uploaded_file = SimpleUploadedFile(
            "test.xlsx",
            excel_content.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        response = self.client.post(
            reverse('upload_resit_details', args=[self.course.id]),
            {'excel_file': uploaded_file},
            format='multipart'
        )
        self.assertEqual(response.status_code, 500)
        self.assertIn('Invalid \'num_questions\' value', response.json()['message'])
