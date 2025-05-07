from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from exams.models import Grade
from users.models import StudentProfile
from .models import Course, ResitExamContent, ResitExamSchedule
from .forms import ExcelUploadForm
from django.contrib.auth.models import Group
import openpyxl
from django.http import JsonResponse
from django.views.decorators.http import require_POST


@login_required
def upload_resit_schedule(request):
    if not request.user.groups.filter(name='faculty').exists():
        return JsonResponse({'status': 'error', 'message': 'Unauthorized access.'}, status=403)

    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                excel_file = request.FILES['excel_file']
                if not excel_file.name.endswith(('.xlsx', '.xls')):
                    return JsonResponse({'status': 'error', 'message': 'Please upload a valid Excel file (.xlsx or .xls).'}, status=400)

                workbook = openpyxl.load_workbook(excel_file)
                worksheet = workbook.active
                headers = [cell.value.lower() if cell.value else '' for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
                expected_headers = ['course id', 'course name', 'place', 'date']

                if not all(header in headers for header in expected_headers):
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Excel file must contain columns: Course ID, Course Name, Place, Date.'
                    }, status=400)

                col_indices = {header: headers.index(header) for header in expected_headers}
                errors = []

                for row_idx, data_row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                    try:
                        course_id = data_row[col_indices['course id']].value
                        course_name = data_row[col_indices['course name']].value
                        place = data_row[col_indices['place']].value
                        date = data_row[col_indices['date']].value

                        try:
                            course_code = str(course_id).strip()
                            course = Course.objects.get(code=course_code)
                            if course_name and str(course_name).strip().lower() != course.name.lower():
                                errors.append(f"Row {row_idx}: Course name '{course_name}' does not match Course ID {course_id}.")
                                continue
                        except (ValueError, TypeError):
                            errors.append(f"Row {row_idx}: Invalid Course ID: {course_id}.")
                            continue
                        except Course.DoesNotExist:
                            errors.append(f"Row {row_idx}: Course with ID {course_id} not found.")
                            continue

                        place = str(place).strip() if place else ''
                        date = str(date).strip() if date else ''
                        if not place or not date:
                            errors.append(f"Row {row_idx}: Place and Date are required.")
                            continue

                        if len(place) + len(date) + 2 > 100:
                            errors.append(f"Row {row_idx}: Combined Place and Date exceed 100 characters.")
                            continue

                        ResitExamSchedule.objects.update_or_create(
                            course=course,
                            defaults={
                                'place': place,
                                'date': date
                            }
                        )
                    except Exception as e:
                        errors.append(f"Row {row_idx}: Error processing data: {str(e)}")

                if errors:
                    return JsonResponse({'status': 'error', 'message': '\n'.join(errors)}, status=400)

                return JsonResponse({'status': 'success', 'message': 'Resit exam schedule uploaded successfully.'})

            except Exception as e:
                return JsonResponse({'status': 'error', 'message': f'Error processing Excel file: {str(e)}'}, status=500)
        else:
            return JsonResponse({'status': 'error', 'message': 'Invalid form submission.'}, status=400)
    else:
        form = ExcelUploadForm()
        return render(request, 'facultysecexam.html', {'form': form})


@require_POST
def upload_resit_details(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    excel_file = request.FILES.get('excel_file')

    if not excel_file:
        return JsonResponse({'status': 'error', 'message': 'No file uploaded'}, status=400)

    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 4:
                raise ValueError(f"Incomplete data in row: {row}")

            num_questions, exam_type, calculator_allowed, additional_notes = row

            if not isinstance(num_questions, (int, float)):
                raise ValueError(f"Invalid 'num_questions' value: {num_questions}")

            calc_allowed = str(calculator_allowed).strip().lower()
            if calc_allowed not in ['yes', 'no']:
                raise ValueError(f"Invalid 'calculator_allowed' value: {calculator_allowed}")
            calculator_bool = calc_allowed == 'yes'

            ResitExamContent.objects.update_or_create(
                course=course,
                defaults={
                    'num_questions': int(num_questions),
                    'exam_type': str(exam_type or '').strip(),
                    'calculator_allowed': calculator_bool,
                    'additional_notes': str(additional_notes or '').strip()
                }
            )

        return JsonResponse({'status': 'success'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required
def resitannouncement(request):
    if not request.user.groups.filter(name='student').exists():
        messages.error(request, "Only students can view resit exam announcements.")
        return redirect('student_page')

    try:
        student_profile = StudentProfile.objects.get(user=request.user)
    except StudentProfile.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('student_page')

    declared_courses = Grade.objects.filter(
        student=student_profile,
        declared_resit=True
    ).values_list('course_id', flat=True)

    schedules = ResitExamSchedule.objects.filter(course__id__in=declared_courses).select_related('course')
    contents = ResitExamContent.objects.filter(course__id__in=declared_courses).select_related('course')

    resit_details = {}
    for sched in schedules:
        resit_details[sched.course.id] = {'schedule': sched}
    for content in contents:
        if content.course.id in resit_details:
            resit_details[content.course.id]['content'] = content
        else:
            resit_details[content.course.id] = {'content': content}

    return render(request, 'resitannouncement.html', {'resit_details': resit_details})





def facultysecexam(request):
    return render(request, 'facultysecexam.html')
